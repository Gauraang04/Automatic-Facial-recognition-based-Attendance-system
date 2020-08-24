from openpyxl import Workbook


def function(n):
    print(n)
    mylist = list(dict.fromkeys(n))
    print(mylist)
    
'''
    w=Workbook()
    sheet=w.active
    print(n)
    while a!=n:
        a=n
        print(n)
    


   
workbook = Workbook()
sheet = workbook.active

sheet["A1"] = "hello"
sheet["B1"] = "world!"

def function2():  # Function to mark attendance in the spreadsheet
    root_2 = Tk()
    root_2.title("ENTER DETAILS")
    root_2.configure(bg="#795548")
    for i in range(0, no_of_stds):
        image.append(face_recognition.load_image_file(
            "database/image"+str(i)+".jpg"))
        encoding.append(face_recognition.face_encodings(image[i])[0])

    def save_attendance():
        wb = xlwt.Workbook()
        ws = wb.add_sheet("My Sheet")
        for i, row in enumerate(DATA):
            for j, col in enumerate(row):
                ws.write(i, j, col)
        wb.save(e2_2.get()+"_"+e1_2.get()+".xls")






workbook.save(filename="hello_world.xlsx")
'''
#function(['Gauraang', 'Gauraang', 'Gauraang', 'Gauraang', 'Gauraang', 'Gauraang', 'Gauraang', 'Gauraang', 'Gauraang', 'Gauraang'])
